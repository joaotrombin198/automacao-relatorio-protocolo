import os
import shutil
from datetime import datetime, timedelta
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.datetime import from_excel
from copy import copy
import xlrd

pasta_destino = r'Q:\Tecnologia da Informacao\Sistemas\Joao_Vitor\teste_protocolos'
nome_base = 'Relatorio-Base.xlsx'
caminho_base = os.path.join(pasta_destino, nome_base)
downloads_dir = os.path.join(os.path.expanduser('~'), 'Downloads')

def encontrar_proxima_linha_vazia(ws, min_row=3):
    for i in range(min_row, ws.max_row + 2):
        # Para cada linha i, verifica se TODAS as células na linha são None ou strings vazias (ou só espaços)
        if all((cell.value is None) or (str(cell.value).strip() == '') for cell in ws[i]):
            return i
    return ws.max_row + 1


def formata_arquivo_novo(wb):
    ws = wb.active

    verde_escuro = "006400"
    fonte_cabecalho = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
    fill_cabecalho = PatternFill(start_color=verde_escuro, end_color=verde_escuro, fill_type="solid")
    alinhamento_cabecalho = Alignment(horizontal="center", vertical="center", wrap_text=True)
    borda_fina = Border(
        left=Side(style='thin', color='FFFFFF'),
        right=Side(style='thin', color='FFFFFF'),
        top=Side(style='thin', color='FFFFFF'),
        bottom=Side(style='thin', color='FFFFFF')
    )

    # Mesclar células
    ws.merge_cells('A1:B1')
    ws.merge_cells('C1:S1')
    ws.merge_cells('T1:T2')

    ws['A1'] = "CONTA"
    ws['C1'] = "PROTOCOLO"
    ws['T1'] = "MOVIMENTAÇÕES"

    conta_cabecalho = ["IDENTIFICADOR", "NOME DA CONTA"]
    protocolo_cabecalho = [
        "DATA DE ABERTURA", "PROTOCOLO", "STATUS", "STATUS GPU", "RESPONSAVEL", "RESPONSAVEL ORIGEM",
        "DATA DE ENCERRAMENTO", "DESCRIÇÃO DE ATENDIMENTO", "DESCRIÇÃO DE ENCERRAMENTO", "DESCRIÇÃO",
        "CATEGORIA", "MOTIVO", "SUBMOTIVO", "CANAL DE ENTRADA", "GRUPO ATENDIMENTO ATUAL",
        "GRUPO ATENDIMENTO ORIGEM", "TIPO DE SOLICITAÇÃO"
    ]

    for i, titulo in enumerate(conta_cabecalho, start=1):
        cell = ws.cell(row=2, column=i, value=titulo)
        cell.font = fonte_cabecalho
        cell.fill = fill_cabecalho
        cell.alignment = alinhamento_cabecalho
        cell.border = borda_fina

    for i, titulo in enumerate(protocolo_cabecalho, start=3):
        cell = ws.cell(row=2, column=i, value=titulo)
        cell.font = fonte_cabecalho
        cell.fill = fill_cabecalho
        cell.alignment = alinhamento_cabecalho
        cell.border = borda_fina
        if i in (3, 7):
            cell.number_format = "DD/MM/YYYY"

    ws['T1'].font = fonte_cabecalho
    ws['T1'].fill = fill_cabecalho
    ws['T1'].alignment = alinhamento_cabecalho
    ws['T1'].border = borda_fina

    for col in range(1, 21):
        for row in [1, 2]:
            cell = ws.cell(row=row, column=col)
            cell.font = fonte_cabecalho
            cell.fill = fill_cabecalho
            cell.alignment = alinhamento_cabecalho
            cell.border = borda_fina

    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 30

    larguras = [20, 25, 15, 30, 20, 20, 20, 20, 25, 20, 25, 20, 20, 25, 20, 25, 20, 25, 30, 40]
    for i, largura in enumerate(larguras, start=1):
        letra_col = chr(64 + i)
        ws.column_dimensions[letra_col].width = largura


def converter_xls_para_xlsx(caminho_xls, caminho_xlsx):
    from openpyxl.utils.datetime import from_excel

    wb_xls = xlrd.open_workbook(caminho_xls)
    sheet_xls = wb_xls.sheet_by_index(0)

    wb_xlsx = Workbook()
    ws_xlsx = wb_xlsx.active

    altura_linha_fina = 15
    colunas_datas = [3, 7]  # colunas 1-based que sao datas

    for row_idx in range(sheet_xls.nrows):
        for col_idx in range(sheet_xls.ncols):
            valor = sheet_xls.cell_value(row_idx, col_idx)
            cell = ws_xlsx.cell(row=row_idx + 1, column=col_idx + 1)

            # Se for coluna de data e valor for número, tenta converter
            if (col_idx + 1) in colunas_datas and isinstance(valor, (int, float)):
                try:
                    cell.value = from_excel(valor)
                    cell.number_format = "DD/MM/YYYY"
                except:
                    cell.value = valor
            else:
                cell.value = valor

        # Força altura da linha fininha
        ws_xlsx.row_dimensions[row_idx + 1].height = altura_linha_fina

    wb_xlsx.save(caminho_xlsx)
    print(f"[OK] Convertido .xls para .xlsx com data corrigida e altura padronizada: {caminho_xlsx}")

def init_xlsx(caminho_arquivo):
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório"

    # Cores e fontes
    verde_escuro = "006400"  # dark green
    fonte_cabecalho = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
    fill_cabecalho = PatternFill(start_color=verde_escuro, end_color=verde_escuro, fill_type="solid")
    alinhamento_cabecalho = Alignment(horizontal="center", vertical="center", wrap_text=True)
    borda_fina = Border(
        left=Side(style='thin', color='FFFFFF'),
        right=Side(style='thin', color='FFFFFF'),
        top=Side(style='thin', color='FFFFFF'),
        bottom=Side(style='thin', color='FFFFFF')
    )

    # Mesclar células para o cabeçalho linha 1
    ws.merge_cells('A1:B1')  # CONTA
    ws.merge_cells('C1:S1')  # PROTOCOLO (colunas C a S)
    ws.merge_cells('T1:T2')  # MOVIMENTAÇÕES (coluna T)

    # Texto cabeçalho linha 1
    ws['A1'] = "CONTA"
    ws['C1'] = "PROTOCOLO"
    ws['T1'] = "MOVIMENTAÇÕES"

    # Cabeçalho linha 2 (detalhado)
    conta_cabecalho = ["IDENTIFICADOR", "NOME DA CONTA"]
    protocolo_cabecalho = [
        "DATA DE ABERTURA", "PROTOCOLO", "STATUS", "STATUS GPU", "RESPONSAVEL", "RESPONSAVEL ORIGEM",
        "DATA DE ENCERRAMENTO", "DESCRIÇÃO DE ATENDIMENTO", "DESCRIÇÃO DE ENCERRAMENTO", "DESCRIÇÃO",
        "CATEGORIA", "MOTIVO", "SUBMOTIVO", "CANAL DE ENTRADA", "GRUPO ATENDIMENTO ATUAL",
        "GRUPO ATENDIMENTO ORIGEM", "TIPO DE SOLICITAÇÃO"
    ]

    # Preencher conta (A2, B2)
    for i, titulo in enumerate(conta_cabecalho, start=1):
        cell = ws.cell(row=2, column=i, value=titulo)
        cell.font = fonte_cabecalho
        cell.fill = fill_cabecalho
        cell.alignment = alinhamento_cabecalho
        cell.border = borda_fina

    # Preencher protocolo (C2 até S2, 17 colunas)
    for i, titulo in enumerate(protocolo_cabecalho, start=3):
        cell = ws.cell(row=2, column=i, value=titulo)
        cell.font = fonte_cabecalho
        cell.fill = fill_cabecalho
        cell.alignment = alinhamento_cabecalho
        cell.border = borda_fina
        # Se for colunas de data, já aplica formatação aqui para segurança
        if i in (3, 7):
            cell.number_format = "DD/MM/YYYY"

    # Preencher MOVIMENTAÇÕES (T1 e T2 já mescladas)
    cell = ws['T1']
    cell.font = fonte_cabecalho
    cell.fill = fill_cabecalho
    cell.alignment = alinhamento_cabecalho
    cell.border = borda_fina

    # Estilo das células mescladas da linha 1 e 2
    for col in range(1, 21):
        for row in [1, 2]:
            cell = ws.cell(row=row, column=col)
            if (row == 1 and (col <= 2 or (3 <= col <= 19) or col == 20)) or row == 2:
                cell.font = fonte_cabecalho
                cell.fill = fill_cabecalho
                cell.alignment = alinhamento_cabecalho
                cell.border = borda_fina

    # Ajustar altura das linhas do cabeçalho
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 30

    # Ajustar largura colunas
    larguras = [20, 25, 15, 30, 20, 20, 20, 20, 25, 20, 25, 20, 20, 25, 20, 25, 20, 25, 30, 40]
    for i, largura in enumerate(larguras, start=1):
        letra_col = chr(64 + i)
        ws.column_dimensions[letra_col].width = largura

    wb.save(caminho_arquivo)
    print(f"[OK] Planilha base inicializada e salva em: {caminho_arquivo}")

# ----------- INICIO DO SCRIPT ------------

arquivos_excel = [f for f in os.listdir(downloads_dir) if f.endswith('.xls') or f.endswith('.xlsx')]
if not arquivos_excel:
    print("[ERRO] Nenhum arquivo Excel (.xls ou .xlsx) encontrado na pasta Downloads.")
    exit()

arquivos_excel_path = [os.path.join(downloads_dir, f) for f in arquivos_excel]
arquivo_novo = max(arquivos_excel_path, key=os.path.getctime)

# Se o arquivo for .xls, converte para .xlsx antes de seguir
if arquivo_novo.lower().endswith('.xls'):
    caminho_convertido = os.path.splitext(arquivo_novo)[0] + '.xlsx'  # 'Relatorio.xls' -> 'Relatorio.xlsx'
    converter_xls_para_xlsx(arquivo_novo, caminho_convertido)
    caminho_para_uso = caminho_convertido
else:
    caminho_para_uso = arquivo_novo

data_ontem = (datetime.now() - timedelta(days=1)).strftime('%Y-%m-%d')

# Colunas que são datas no relatório (1-based)
colunas_datas = [3, 7]

altura_linha_fina = 15

if not os.path.exists(caminho_base):
    # Cenário 1: base não existe, cria base formatada e copia dados do arquivo convertido

    init_xlsx(caminho_base)

    wb_base = load_workbook(caminho_base)
    ws_base = wb_base.active

    wb_novo = load_workbook(caminho_para_uso)
    ws_novo = wb_novo.active

    # Copia dados a partir da linha 3 para não sobrescrever cabeçalho mesclado
    for i, row in enumerate(ws_novo.iter_rows(min_row=3, values_only=True), start=3):
        for col_idx, valor in enumerate(row, start=1):
            # Convertendo número de data para datetime se necessário
            if col_idx in colunas_datas and isinstance(valor, (int, float)):
                try:
                    valor = from_excel(valor)
                except:
                    pass
            ws_base.cell(row=i, column=col_idx).value = valor
            # Força formato de data nas colunas
            if col_idx in colunas_datas:
                ws_base.cell(row=i, column=col_idx).number_format = "DD/MM/YYYY"
        # Força altura fina da linha
        ws_base.row_dimensions[i].height = altura_linha_fina

    wb_base.save(caminho_base)
    print(f"[OK] Arquivo base criado com dados do arquivo convertido: {caminho_base}")

else:
    # Cenário 2: base existe, move novo arquivo e cola dados novos na base
    nome_arquivo_dia = f"Relatorio-{data_ontem}.xlsx"
    caminho_destino = os.path.join(pasta_destino, nome_arquivo_dia)
    shutil.move(caminho_para_uso, caminho_destino)
    print(f"[OK] Novo relatório movido para: {caminho_destino}")

    wb_base = load_workbook(caminho_base)
    ws_base = wb_base.active

    wb_novo = load_workbook(caminho_destino)

    # 🔧 Aplica a mesma formatação que usamos na base
    formata_arquivo_novo(wb_novo)
    wb_novo.save(caminho_destino)

    ws_novo = wb_novo.active

    def copiar_estilo(origem_celula, destino_celula):
        destino_celula.font = copy(origem_celula.font)
        destino_celula.border = copy(origem_celula.border)
        destino_celula.fill = copy(origem_celula.fill)
        destino_celula.number_format = origem_celula.number_format
        destino_celula.protection = copy(origem_celula.protection)
        destino_celula.alignment = copy(origem_celula.alignment)

       # Linha modelo para copiar estilos (última linha já existente, assumindo que sempre tem dados)
    linha_modelo = ws_base.max_row
    linha_destino = encontrar_proxima_linha_vazia(ws_base, min_row=3)

    # Copia dados do novo arquivo a partir da linha 3 (pulando cabeçalho)
    for i, row in enumerate(ws_novo.iter_rows(min_row=3), start=linha_destino):
        for col_idx, celula_nova in enumerate(row, start=1):
            valor = celula_nova.value
            # Convertendo número de data para datetime se necessário
            if col_idx in colunas_datas and isinstance(valor, (int, float)):
                try:
                    valor = from_excel(valor)
                except:
                    pass

            celula_base = ws_base.cell(row=i, column=col_idx)
            celula_base.value = valor

            # Força formato de data nas colunas
            if col_idx in colunas_datas:
                celula_base.number_format = "DD/MM/YYYY"

            celula_modelo = ws_base.cell(row=linha_modelo, column=col_idx)
            copiar_estilo(celula_modelo, celula_base)

        # Ajustar altura da linha igual à linha modelo, ou 15 se modelo não definir altura
        altura_modelo = ws_base.row_dimensions[linha_modelo].height
        ws_base.row_dimensions[i].height = altura_modelo if altura_modelo else altura_linha_fina

    wb_base.save(caminho_base)
    print(f"[OK] Dados do relatório novo colados no arquivo base, mantendo formatação e cabeçalho.")